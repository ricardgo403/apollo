# -*- coding: utf-8 -*-
"""
Created on Fri Sep 24 17:07:33 2021

@author: Juan Antonio

Title: funciones de movimiento

ver: 2.0
"""

# Librerias
import cv2 as cv
import numpy as np
import matplotlib.pyplot as plt
import openpyxl as op

# Nota
"""
Este codigo esta creado en un sistema de 480 x 640,
se recomienda adaptar frames a tomar a estas dimensiones 
para evitar perdida de datos.

"""


# Funciones

def union(a, b):
    '''
    Union de rectangulos.\n
    \n
    Parameters\n
    ----------\n
    a : rectangulo 1 [x,y,w,h].\n
    b : rectangulo 2 [x,y,w,h].\n
    \n
    Returns\n
    -------\n
    x.\n
    y.\n
    w.\n
    h.\n

    '''
    x = min(a[0], b[0])
    y = min(a[1], b[1])
    w = max(a[0] + a[2], b[0] + b[2]) - x
    h = max(a[1] + a[3], b[1] + b[3]) - y
    return (x, y, w, h)


def contmax(contornos):
    may = cv.contourArea(contornos[0])
    n = 0
    num = 0
    for c in contornos:
        loc = cv.contourArea(c)
        if loc > may:
            may = loc
            num = n
        n += 1
    return contornos[num]


def d2p(x1, y1, x2, y2):
    d = ((x2 - x1) ** 2 + (y2 - y1) ** 2) ** 0.5

    return d


def ruido(contornosd, contlistd, radio):
    for c in contornosd:
        if cv.contourArea(c) > (3.14 * radio ** 2) / 4:
            contlistd.append(c)

            # --- Rectangulo mayor --- #
    recmax = cv.boundingRect(contlistd[0])
    for c in contlistd:
        recmax = union(recmax, cv.boundingRect(c))
    return recmax, contlistd


def profundidad(frame, tono):
    gframe = cv.cvtColor(frame, cv.COLOR_BGR2GRAY)
    blancos = cv.threshold(gframe, tono, 255, cv.THRESH_BINARY)[1]
    try:
        cont, _ = cv.findContours(blancos, cv.RETR_TREE, cv.CHAIN_APPROX_SIMPLE)
        mayor = contmax(cont)
        x, y, h, w = cv.boundingRect(mayor)
        a = x
        b = y + w
        cutted = blancos[y + int(h / 15):y + w - int(h / 3.5), x + int(h / 5): x + h - int(h / 4)]
        inv = cv.bitwise_not(cutted)
        continv, _ = cv.findContours(inv, cv.RETR_TREE, cv.CHAIN_APPROX_SIMPLE)
        mayorinv = contmax(continv)

        x, y, h, w = cv.boundingRect(mayorinv)
        cm5 = d2p(x, y, x + h, y)
        cm1 = cm5 / 5
        cords = [a, b - int(cm1 * 7)]
        cv.imshow('inv', inv)
        return cm1, inv, cords, True
    except:
        return None, None, None, False


def movimiento(frame, gframe1, blur=5, sensibilidad_morph=10, sensibilidad_dilation=20):
    '''
    Actualización de variables para metodos externos morph y dilation.\n
    \n
    Parameters\n
    ----------\n
    frame : frame actual.\n
    gframe1 : frame anterior.\n
    blur : valor de blur impar. The default is 5.\n
    sensibilidad_morph : valor de uno minimo 0-255. The default is 10.\n
    sensibilidad_dilation : valor de uno minimo 0-255. The default is 20.\n
    \n
    Returns\n
    -------\n
    umbralm: umbral para metodo morph.\n
    umbrald: umbral para metodo dilation.\n
    framedif: diferencia frame pasado y actual.\n
    gframe1: frame actual para futuro cambio.\n

    '''

    bframe = cv.medianBlur(frame, blur)
    grayframe = cv.cvtColor(bframe, cv.COLOR_BGR2GRAY)
    framedif = cv.absdiff(grayframe, gframe1)
    gframe1 = grayframe
    umbralm = cv.threshold(framedif, sensibilidad_morph, 255, cv.THRESH_BINARY)[1]
    umbrald = cv.threshold(framedif, sensibilidad_dilation, 255, cv.THRESH_BINARY)[1]

    return umbralm, umbrald, framedif, gframe1


def morphdata(umbralm, radio, blank=np.zeros((480, 640), dtype='uint8'),
              zeros=np.zeros((480, 640), dtype='uint8'),
              kernelm=np.ones((8, 8), np.uint8)):
    blankc = blank.copy()
    # Funcion
    if not (umbralm == zeros).all():
        morph = cv.morphologyEx(umbralm, cv.MORPH_OPEN, kernelm)

        contornosm, _ = cv.findContours(morph, cv.RETR_TREE, cv.CHAIN_APPROX_SIMPLE)
        try:
            mayor = contmax(contornosm)
            x, y, h, w = cv.boundingRect(mayor)

            xa = x - (5 * radio)
            ya = y - (2 * radio)
            wa = w + (10 * radio)
            ha = h + (4 * radio)

            cv.rectangle(blankc, (xa, ya), (xa + wa, ya + ha), 255, -1)
            return blankc, True
        except:

            return None, False
    else:
        return None, False


def dilationdata(umbrald, mascara, radio, blank=np.zeros((480, 640), dtype='uint8'),
                 cir=np.zeros((480, 640, 3), dtype='uint8'), zeros=np.zeros((480, 640), dtype='uint8')):
    '''
    Seguimiento de movimiento con metodo dilation.\n
    \n
    Parameters\n
    ----------\n
    umbrald : matriz con datos para dilation.\n
    mascara : lugar a aplicar metodo.\n
    radio : radio en pixeles.\n
    blank : matriz a aplicar mascara con contornos. The default is np.zeros((480,640),dtype='uint8').\n
    cir : matriz con ubicacion de circulos. The default is np.zeros((480,640,3),dtype='uint8').\n
    zeros : matriz ceros estatica. The default is np.zeros((480,640),dtype='uint8').\n
    \n
    Returns\n
    -------\n
    blank3c: matriz con contornos y mascara.\n
    circ: matriz con circulos de movimiento.\n
    cm: circulo medio.\n
    ci: circulo izquierdo.\n
    cd: circulo derecho.\n
    True --- False\n

    '''
    # Locales
    blank3c = blank.copy()
    contlistd = list()
    maski = np.zeros((480, 640), dtype='uint8')
    maskd = np.zeros((480, 640), dtype='uint8')
    circ = cir.copy()
    # Funcion
    if not (umbrald == zeros).all():
        dilation = cv.dilate(umbrald, None, iterations=1)
        try:
            masked = cv.bitwise_and(dilation, dilation, mask=mascara)

            # --- Eliminacion de contornos con area menor a areacir/4 --- #
            contornosd, hierarchyd = cv.findContours(masked, cv.RETR_TREE, cv.CHAIN_APPROX_SIMPLE)

            recmax, contlistd = ruido(contornosd, contlistd, radio)

            cv.drawContours(blank3c, contlistd, -1, 255, -1)
            cv.rectangle(blank3c, (recmax[0], recmax[1]), (recmax[0] + recmax[2], recmax[1] + recmax[3]), (0, 255, 0),
                         1)

            # --- Separación iz y der maskmethod--- #
            cv.rectangle(maski, (recmax[0], recmax[1]), (recmax[0] + int(recmax[2] / 4), recmax[1] + recmax[3]), 255,
                         -1)
            maski = cv.bitwise_and(blank3c, blank3c, mask=maski)

            cv.rectangle(maskd, (recmax[0] + int(recmax[2] * (3 / 4)), recmax[1]),
                         (recmax[0] + recmax[2], recmax[1] + recmax[3]), 255, -1)
            maskd = cv.bitwise_and(blank3c, blank3c, mask=maskd)

            # --- Punto mas alto de iz y der --- #
            contornosdi, hierarchydi = cv.findContours(maski, cv.RETR_TREE, cv.CHAIN_APPROX_SIMPLE)
            reci = cv.boundingRect(contornosdi[0])
            for c in contornosdi:
                reci = union(reci, cv.boundingRect(c))
            contornosdd, hierarchydd = cv.findContours(maskd, cv.RETR_TREE, cv.CHAIN_APPROX_SIMPLE)
            recd = cv.boundingRect(contornosdd[0])
            for c in contornosdd:
                recd = union(recd, cv.boundingRect(c))
                # --- Creacion de circulos iz y der --- #
            ci = [(reci[0] + radio, reci[1] + radio), radio]

            cd = [(recd[0] - radio + recd[2], recd[1] + radio), radio]

            cv.circle(circ, ci[0], ci[1], (255, 0, 0), 1)
            cv.circle(circ, cd[0], cd[1], (0, 0, 255), 1)

            # --- Creacion de circulo medio --- #
            cm = [(int((ci[0][0] + cd[0][0]) / 2), int((ci[0][1] + cd[0][1]) / 2)), radio]

            cv.circle(circ, cm[0], cm[1], (0, 255, 0), -1)
            # --- Paso de cordenadas posicionales de circulo y tiempo entre toma --- #
            return blank3c, circ, cm, ci, cd, True
        except:
            return False, False, False, False, False, False
    else:
        return False, False, False, False, False, False


def imshows(*pantallas):
    '''
    Actualizar frames. \n
    \n
    Parameters\n
    ----------\n
    *pantallas : nombres de las variables a actualizar.\n
    \n
    Returns\n
    -------\n
    None.\n

    '''
    n = 0
    for pantalla in pantallas:
        cv.imshow(f'{n}', pantalla)
        n += 1


def division_up_down(obj, pixeles=140):
    '''
    Division por secciones arriba y abajo\n
    \n
    Parameters\n
    ----------\n
    obj : matriz a modificar.\n
    pixeles : pixeles de punto de cambio. The default is 140.\n
    \n
    Returns\n
    -------\n
    up : matriz parte 1.\n
    down : matriz parte 2.\n

    '''

    blanku = np.zeros((480, 640), dtype='uint8')
    cv.rectangle(blanku, (0, 0), (640, pixeles), 255, -1)
    up = cv.bitwise_and(obj, obj, mask=blanku)

    blankd = np.zeros((480, 640), dtype='uint8')
    cv.rectangle(blankd, (0, pixeles), (640, 480), 255, -1)
    down = cv.bitwise_and(obj, obj, mask=blankd)

    return up, down


def data(datav, *valor):
    loc = list()
    for val in valor:
        loc.append([val[0], val[1]])
    datav.append(loc)
    return datav


def inclinacion(frameup, gframeup1, radio, radpx, cords, kernelm=np.ones((1, 1), np.uint8)):
    blank = np.zeros((480, 640, 3), dtype='uint8')
    cv.circle(blank, (cords[0], cords[1]), radpx, (255, 0, 0), 1)

    gframeup = cv.cvtColor(frameup, cv.COLOR_BGR2GRAY)
    dif = cv.absdiff(gframeup, gframeup1)
    cv.imshow('dif', dif)
    umbral = cv.threshold(dif, 10, 255, cv.THRESH_BINARY)[1]
    cv.imshow('umb', umbral)
    morph = cv.morphologyEx(umbral, cv.MORPH_OPEN, kernelm)
    try:
        cont, _ = cv.findContours(morph, cv.RETR_TREE, cv.CHAIN_APPROX_SIMPLE)
        contlistd = list()
        recmax, contlistd = ruido(cont, contlistd, radio)
        cv.drawContours(blank, contlistd, -1, (0, 0, 255), -1)
        cv.rectangle(blank, (recmax[0], recmax[1]), (recmax[0] + recmax[2], recmax[1] + recmax[3]), (0, 255, 0), 1)
        cv.imshow('rec', blank)
        """
        p1x = (radpx**2 - (recmax[1]-cords[1])**2)**0.5
        
        p1 = [p1x + cords[0] , recmax[1]]
        p2y = (radpx**2 - (recmax[0]-cords[0])**2)**0.5
        
        p2 = [recmax[0], p2y +cords[1]]
        """

        p1 = [recmax[0], recmax[1] + recmax[3]]
        cv.imshow('c', blank)
        print('end')
        return p1, True
    except:
        print(Exception)
        return None, False


def graph(xa, ya, nombre):
    fig, ax = plt.subplots()
    ax.plot(xa, ya)
    ax.scatter(xa, ya, marker='>')
    plt.savefig(f'{nombre}.png')
    plt.show()
    return plt


def graphfromsublist(lista, index, nombre):
    x = list()
    y = list()
    for el in lista:
        x.append(el[index][1][0])
        y.append(el[index][1][1])
    fig, ax = plt.subplots()
    ax.plot(x, y)
    ax.scatter(x, y, marker='>')
    plt.savefig(f'{nombre}.png')
    plt.show()
    return plt


def excel(lista, nombre):
    print('Excel abierto')
    wb = op.Workbook()
    hoja = wb.active
    hoja.title = 'data'

    titles = len(lista[0])
    for x in range(titles):
        hoja.cell(row=1, column=x + 1, value=lista[0][x][0])
    y = 2
    for el in lista:
        x = 1
        for e in el:
            var = ''
            if isinstance(e[1], list):
                var = str(e[1])
            else:
                var = e[1]
            hoja.cell(row=y, column=x, value=var)
            x += 1
        y += 1
    wb.save(f'{nombre}.xlsx')
    print(f'Excel guardado como: {nombre}.xlsx')
